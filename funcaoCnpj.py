#Projeto de consulta de CNPJ

import time
import requests
import json
import openpyxl
import os

#Planilha que retorna os resultados
arquivo = "C:/Users/welli/OneDrive/Área de Trabalho/consultaCnpj/pesquisa_cnpj.xlsx"

#Instanciando uma planilha de openpyxl
book = openpyxl.Workbook()

#Utilizando o objeto book para criar uma nova tabela dentro da planilha
book.create_sheet("cnpj_pesquisados")

#Guardando dentro de uma variavel a minha tabela
cnpj_page=book["cnpj_pesquisados"]

#Verificando se a minha planilha de pesquisa de cnpj existe...
if os.path.isfile(arquivo):
   #carrego os dados da planilha dentro da variavel planilha_cnpj
   planilha_cnpj = openpyxl.load_workbook("cnpj_para_pesquisar.xlsx")

   #entrando na tabela dentro da minha planilha
   sheet_page = planilha_cnpj["Sheet1"]
   
   #criando um array de cnpj
   cnpj = []

   #Para cada linha dentro da tabela sheet1
   for rows in sheet_page.iter_rows(max_col=1, min_row=2):
      #para cada celula nas linhas
      for cell in rows:

         if cell.value != None:
            cnpj_formatado = cell.value
            cnpj_formatado = cnpj_formatado.replace(".","")
            cnpj_formatado = cnpj_formatado.replace("/","")
            cnpj_formatado = cnpj_formatado.replace("-","")
            #adicionando dentro do array de cnpj o cnpj formatado
            cnpj.append({"cnpj": cnpj_formatado})

#Guardando dentro de uma variavel o ultimo valor do array cnpj
ultimo_cnpj = cnpj[-1]

print("O programa irá demorar alguns minutos para finalizar... Por favor aguarde!")

#Adicionando um cabeçalho
cnpj_page.append(["CNPJ",  "NOME", "SITUACAO", "DATA_SITUACAO"])

#funçao que consulta a API da receita
def consulta_cnpj(cnpj):
   print("Consultando CNPJ:" +cnpj)
   url = f"https://receitaws.com.br/v1/cnpj/{cnpj}"
   response = requests.request("GET", url)

   resp = json.loads(response.text)
   if resp['status'] == 'ERROR':
      print("CNPJ invalido: " +cnpj+ "... Prosseguindo para a proxima consulta...")
      return
   else:
      print("Sucesso... Prosseguindo para a proxima consulta...")

   #Dados da API
   nome = (resp['nome'])
   situacao = (resp['situacao'])
   data_situacao = (resp['data_situacao'])
   cnpj_empresa = (resp['cnpj'])
   
   cnpj_page.append([cnpj_empresa,  nome, situacao, data_situacao])

   
#para Cada item dentro de CPNJ
for item in cnpj:
   consulta_cnpj(item["cnpj"]) 

   #se o cpnj for o ultimo ele finaliza a execuçao
   if (item==ultimo_cnpj):
      print("Programa finalizado com sucesso!")
      book.save("pesquisa_cnpj.xlsx")
      exit()
   time.sleep(20)

























